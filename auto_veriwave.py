# Veriwave Test Process
# connect to 2.4G or 5G card to CPE
# connect console to CPE (COM1) and enter bash mode
# python to call vw_auto.tcl (tclsh vw_auto.tcl) with variable
# variable MCS -> Spatial Stream -> protocol -> direction
# change channel (36/64/100/149/157) and repeat the above test
#                   "tclsh84", VWTEV.VW_TEST_ROOT + "\\bin\\vw_auto.tcl",
#                   "--savepcaps",
#                   "-f", VWTEV.VW_TEST_ROOT + "\\conf\\" + "CTL.tcl",
#                   "-i", strResultDirName,
#                   "--noconfig", "--tests", strTest,
#                   "--debug", Integer.toString(vwtev.intvw_autoDebug),
#                   "--var", "mAddr", VWTEV.VWAddrChassis,
#                   "--var", "mClients", Integer.toString(intClientCnt % 100),
#                   "--var", "mEClients_1", Integer.toString(intEClientCnt_1),
#                   "--var", "mEClients_2", Integer.toString(intEClientCnt_2),
#                   "--var", "mFrSz", Integer.toString(intFrameSize),
#                   "--var", "mDir", strDirection,
#                   "--var", "mSrc", strSource,
#                   "--var", "mDst", strDestination,
#
# #$ip_addr	Veriwave IP address	$ip_addr
# $ac_port	Veriwave AC port	2:1
# $n_port	Veriwave n port	2:1
# $eth_port	Veriwave ETH port	1:1
# $logdir	Log folder	xxxxx
# $direction	Traffic direction	Unidirectional, Bidirectional
# $channel_ac	802.11ac channel	36, 64, 100, 149, 157
# $channel_n	802.11n channel	1, 6, 11
# $client_num	client number	1~20
# $src	Source group	dut_eth, dut_ac, dut_n
# $dst	Destination group	dut_eth, dut_ac, dut_n
# $traffic	Traffic type	Udp,Tcp,Icmp,Raw
# $stream	Spatial Stream	1,2,3,4
# $mcs_ac	802.11ac MCS	9,8,7,0
# $mcs_n	802.11n MCS	11,31
# $bandwidth_n 802.11n bandwidth 20,40
# $frame	Frame Size	128 256 512 1024 1280 1518
# $version	CPE FW version	xxxxx
# $model	CPE model name	xxxxx
#
# tclsh84 vw_auto.tcl --savepcaps -i c:\\temp --noconfig --tests TPut --debug 0 --var ip_addr 10.1.1.180 ^
# --var ac_port	2:1 --var n_port 2:1 --var eth_port 1:1 --var logdir c:\\temp --var direction Unidirectional ^
# --var channel_ac 36 --var channel_n 1 --var client_num 1 --var src dut_eth --var dst dut_ac ^
# --var traffic	Udp --var stream 4 --var mcs_ac 9 --var mcs_n 31 ^
# --var frame "128 256 512 1024 1280 1518" --var version b1 --var model CPE

# !/usr/bin/env python

import os
import json
import sys
import csv
import tkinter as tk
import tkinter.filedialog as fd
import tkinter.messagebox as msgbox
from tkinter import ttk
import re
import openpyxl
#from viusalkey import *
import time
#import win32gui
import win32con
#import win32api
from ctypes import *
import ctypes
import ctypes.wintypes
import subprocess
from threading import Thread
import shlex
from shutil import copyfile
import logging
import io
from datetime import date

# for ssh function
# import paramiko, base64
# import datetime

# logging setting,10=DEBUG,20=INFO
output_logging = 20

#For Veriwave site
QTN_CONSOLE = "COM6:115200baud - Tera Term VT"
CPE_CONSOLE = "COM1:115200baud - Tera Term VT"
vQTN_Console = "COM6:115200baud - Tera Term VT"
vCPE_Console = "COM1:115200baud - Tera Term VT"

#For Testing
#QTN_CONSOLE = "COM1 - Tera Term VT"
#CPE_CONSOLE = "COM1 - Tera Term VT"

CHASSIS_IP_ADDR = "10.1.1.180"
vChassis_ip_addr = "10.1.1.180"
LICENSE_IP_ADDR = "10.1.1.180"
vLicense_ip_addr = "10.1.1.180"
PORT_AC = "2:1"
vPort_AC = "2:1"
PORT_N = "2:1"
vPort_N = "2:1"
PORT_ETH = "1:1"
vPort_ETH = "1:1"
LOG_BASE_DIR = "C:\\temp"
vLog_Dir = ""
TEST_TYPE = ["802.11ax_2.4G", "802.11ax_5G", "802.11ac", "802.11n"]
vTest_Type = "802.11ac"
TEST_PROFILE = ["Full", "Single"]
vTest_Profile = "Full"
VERSION = "v1.00(AAA.0)b1"
vVersion = "v1.00(AAA.0)b1"
MODEL = "CPE"
vModel = "CPE"
DIRECTION = ["Unidirectional", "Bidirectional"]
vDirection = "Unidirectional,Bidirectional"
FLOW_DIR = ["TX", "RX"]
CHANNEL_AC = ["36", "64", "100"]
vChannel_AC = "36,64,100"
CHANNEL_N = ["1", "6", "11"]
vChannel_N = "1,6,11"
CLIENT_NUM = "1"
vClient_Num = "1"
ANTENNA_NUM = ["4", "2"]
vAntenna_Num = "4"
BANDWIDTH_N = ["20", "40"]
vBandwidth_n = "20"
TRAFFIC_SRC = ["dut_eth", "dut_ac", "dut_n"]
vTraffic_Src = "dut_eth"
TRAFFIC_DST = ["dut_eth", "dut_ac", "dut_n"]
vTraffic_Dst = "dut_ac"
TRAFFIC_TYPE = ["Udp", "Tcp"]
vTraffic_Type = "Udp"
STREAM = ["4", "3", "2", "1"]
vStream = "4"
MCS_AC = ["9", "8", "7", "0"]
vMcs_ac = "9"
MCS_N = ["31","23","15","7"]
vMcs_n = "31"
FRAME = ["128", "256", "512", "1024", "1280", "1518"]
vFrame = "128,256,512,1024,1280,1518"
PROJECT_TYPE = ["HGU","VB"]
vProject_Type = "HGU"

report_generate = True

AC_START_COL = 5
AC_START_ROW = 10
AC_TEF_LOC_COL = 2
AC_REF_LOC_COL = 4
AC_PKT_SIZE_ROW = 6
AC_VAL_SIZE_COL = 6
AC_ATN_SIZE = 4

N_START_COL = 7
N_START_ROW = 4
N_TEF_LOC_COL = 0
N_REF_LOC_COL = -1
N_PKT_SIZE_ROW = 6
N_VAL_SIZE_COL = 4
N_ATN_SIZE = 4

ac_data_loc = {
    'UDP_DOWN': 0,
    'UDP_UP': 1 * AC_PKT_SIZE_ROW,
    'TCP_DOWN': 2 * AC_PKT_SIZE_ROW,
    'TCP_UP': 3 * AC_PKT_SIZE_ROW,
    'UDP_BI': 4 * AC_PKT_SIZE_ROW,
    'CANAL36': AC_START_ROW,
    'CANAL64': AC_START_ROW + AC_PKT_SIZE_ROW * 5,
    'CANAL100': AC_START_ROW + AC_PKT_SIZE_ROW * 10,
    'CANAL149': AC_START_ROW + AC_PKT_SIZE_ROW * 15,
    'CANAL157': AC_START_ROW + AC_PKT_SIZE_ROW * 20,
    '128': 0,
    '256': 1,
    '512': 2,
    '1024': 3,
    '1280': 4,
    '1518': 5,
    '4x4': 0,
    '3x3': 1 * AC_VAL_SIZE_COL,
    '2x2': 2 * AC_VAL_SIZE_COL,
    '1x1': 3 * AC_VAL_SIZE_COL,
    'MCS_9': AC_START_COL,
    'MCS_8': AC_START_COL + AC_VAL_SIZE_COL * AC_ATN_SIZE * 1 + 1,
    'MCS_7': AC_START_COL + AC_VAL_SIZE_COL * AC_ATN_SIZE * 2 + 2,
    'MCS_0': AC_START_COL + AC_VAL_SIZE_COL * AC_ATN_SIZE * 3 + 3
}

n_data_loc = {
    'UDP_DOWN': 0,
    'UDP_UP': 1 * N_PKT_SIZE_ROW,
    'TCP_DOWN': 2 * N_PKT_SIZE_ROW,
    'TCP_UP': 3 * N_PKT_SIZE_ROW,
    'UDP_BI': 4 * N_PKT_SIZE_ROW,
    'CANAL1': N_START_ROW,
    'CANAL6': N_START_ROW + N_PKT_SIZE_ROW * 5,
    'CANAL11': N_START_ROW + N_PKT_SIZE_ROW * 10,
    'CANAL13': N_START_ROW + N_PKT_SIZE_ROW * 15,
    'BW20': N_START_COL,
    'BW40': N_START_COL + N_VAL_SIZE_COL * N_ATN_SIZE * 1,
    '128': 0,
    '256': 1,
    '512': 2,
    '1024': 3,
    '1280': 4,
    '1518': 5,
    '4x4': 0,
    '3x3': 1 * N_VAL_SIZE_COL,
    '2x2': 2 * N_VAL_SIZE_COL,
    '1x1': 3 * N_VAL_SIZE_COL,
}

# v1.0 add 11ax and 11ac test sheet for 5G
# v1.0 add report_name with date info

class Auto_Veriwave(tk.Tk):
    def __init__(self, top=None):
        super().__init__()
        self.thread = 0
        self.proc_state = False
        x_loc = 0.01
        y_loc = 0.01
        x_gap = 0.2
        y_gap = 0.05

        # v1.0 initial release
        # v1.1 add 11ax_2.4g profile testing
        # v1.11 add MSC9/0 checking in the folder name when utility generate the report
        
        self.start_logging()
        self.geometry("830x900+250+50")
        self.title("Veriwave Automation v1.11 (2022/3/21)")

        self.ChassisIP_Label = tk.Label(self, text='Chassis IP address:')
        self.ChassisIP_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.ChassisIP = tk.Entry(self, width=100)
        self.ChassisIP.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.ChassisIP.insert(tk.END, CHASSIS_IP_ADDR)

        y_loc += y_gap
        self.LicenseIP_Label = tk.Label(self, text='License IP address:')
        self.LicenseIP_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.LicenseIP = tk.Entry(self, width=100)
        self.LicenseIP.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.LicenseIP.insert(tk.END, LICENSE_IP_ADDR)

        y_loc += y_gap
        self.AC_Port_Label = tk.Label(self, text='802.11ac Port:')
        self.AC_Port_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.AC_Port = tk.Entry(self, width=100)
        self.AC_Port.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.AC_Port.insert(tk.END, PORT_AC)

        y_loc += y_gap
        self.N_Port_Label = tk.Label(self, text='802.11n Port:')
        self.N_Port_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.N_Port = tk.Entry(self, width=100)
        self.N_Port.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.N_Port.insert(tk.END, PORT_N)

        y_loc += y_gap
        self.EthPort_Label = tk.Label(self, text='Ethernet Port:')
        self.EthPort_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.EthPort = tk.Entry(self, width=100)
        self.EthPort.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.EthPort.insert(tk.END, PORT_ETH)

        y_loc += y_gap
        self.LogDir_Label = tk.Label(self, text='Log Directory:')
        self.LogDir_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.LogDir = tk.Entry(self, width=100)
        self.LogDir.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.LogDir.insert(tk.END, LOG_BASE_DIR)

        y_loc += y_gap
        self.Model_Label = tk.Label(self, text='CPE Model:')
        self.Model_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.Model = tk.Entry(self, width=100)
        self.Model.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.Model.insert(tk.END, MODEL)

        y_loc += y_gap
        self.Version_Label = tk.Label(self, text='CPE Version:')
        self.Version_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.Version = tk.Entry(self, width=100)
        self.Version.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.Version.insert(tk.END, VERSION)

        y_loc += y_gap
        self.TestType_Label = tk.Label(self, text='Test Type:')
        self.TestType_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.TestType = ttk.Combobox(self, width=100, values=TEST_TYPE)
        self.TestType.current(0)
        self.TestType.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.TestType.bind("<<ComboboxSelected>>", self.test_profile_change)

        y_loc += y_gap
        self.AntennaNum_Label = tk.Label(self, text='CPE Antenna:')
        self.AntennaNum_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.AntennaNum = ttk.Combobox(self, width=100, values=ANTENNA_NUM)
        self.AntennaNum.current(0)
        self.AntennaNum.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.AntennaNum.bind("<<ComboboxSelected>>", self.antenna_change)

        y_loc += y_gap
        self.TestProfile_Label = tk.Label(self, text='Test Profile:')
        self.TestProfile_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.TestProfile = ttk.Combobox(self, width=100, values=TEST_PROFILE)
        self.TestProfile.bind("<<ComboboxSelected>>", self.test_profile_change)
        self.TestProfile.current(0)
        self.TestProfile.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)

        y_loc += y_gap
        self.ProjectType_Label = tk.Label(self, text='Project Type:')
        self.ProjectType_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.ProjectType = ttk.Combobox(self, width=100, values=PROJECT_TYPE)
        self.ProjectType.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.ProjectType.current(0)

        # Optional Selection
        x_loc = x_loc + 0.45
        y_loc = 0.01
        self.ClientNum_Label = tk.Label(self, text='Client Number:')
        self.ClientNum_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.ClientNum = tk.Entry(self, width=100)
        self.ClientNum.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.ClientNum.insert(tk.END, CLIENT_NUM)

        y_loc += y_gap
        self.N_ChannelList_Label = tk.Label(self, text='11n Channel List:')
        self.N_ChannelList_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.N_ChannelList = tk.Entry(self, width=100)
        self.N_ChannelList.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.N_ChannelList.insert(tk.END, ",".join(CHANNEL_N))

        y_loc += y_gap
        self.N_MCS_Label = tk.Label(self, text='11n MCS:')
        self.N_MCS_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.N_MCS = tk.Entry(self, width=100)
        self.N_MCS.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.N_MCS.insert(tk.END, ",".join(MCS_N))

        y_loc += y_gap
        self.N_Bandwidth_Label = tk.Label(self, text='11n Bandwidth:')
        self.N_Bandwidth_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.N_Bandwidth = tk.Entry(self, width=100)
        self.N_Bandwidth.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.N_Bandwidth.insert(tk.END, ",".join(BANDWIDTH_N))

        y_loc += y_gap
        self.AC_ChannelList_Label = tk.Label(self, text='11ac/ax Channel List:')
        self.AC_ChannelList_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.AC_ChannelList = tk.Entry(self, width=100)
        self.AC_ChannelList.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.AC_ChannelList.insert(tk.END, ",".join(CHANNEL_AC))

        y_loc += y_gap
        self.AC_MCS_Label = tk.Label(self, text='11ac/ax MCS:')
        self.AC_MCS_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.AC_MCS = tk.Entry(self, width=100)
        self.AC_MCS.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.AC_MCS.insert(tk.END, ",".join(MCS_AC))

        y_loc += y_gap
        self.SpatialStream_Label = tk.Label(self, text='Spatial Streams:')
        self.SpatialStream_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.SpatialStream = tk.Entry(self, width=100)
        self.SpatialStream.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.SpatialStream.insert(tk.END, ",".join(STREAM))

        y_loc += y_gap
        self.Direction_Label = tk.Label(self, text='Direction:')
        self.Direction_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.Direction = tk.Entry(self, width=100)
        self.Direction.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.Direction.insert(tk.END, ",".join(DIRECTION))

        y_loc += y_gap
        self.TrafficType_Label = tk.Label(self, text='Traffic Type:')
        self.TrafficType_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.TrafficType = tk.Entry(self, width=100)
        self.TrafficType.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.TrafficType.insert(tk.END, ",".join(TRAFFIC_TYPE))

        y_loc += y_gap
        self.FrameSize_Label = tk.Label(self, text='Frame Size:')
        self.FrameSize_Label.place(relx=x_loc, rely=y_loc, height=32, width=150)
        self.FrameSize = tk.Entry(self, width=100)
        self.FrameSize.place(relx=x_loc + x_gap, rely=y_loc, relheight=0.03, relwidth=0.25)
        self.FrameSize.insert(tk.END, ",".join(FRAME))

        y_loc += y_gap
        self.GenerateButton = tk.Button(self, pady="0", text='Generate Report', command=self.generate)
        self.GenerateButton.place(relx=x_loc + 0.05, rely=y_loc, height=31, width=150)
        self.StartButton = tk.Button(self, pady="0", text='Start Test', command=self.start_test)
        self.StartButton.place(relx=x_loc + 0.3, rely=y_loc, height=31, width=150)

        x_loc = 0.01
        y_loc = 0.6
        # Configure the scrollbars
        self.ResponseText = tk.Text(self, font=("Helvetica", 8))
        self.ResponseText.place(relx=x_loc, rely=y_loc, height=380, width=800)
        self.ScrollBar = tk.Scrollbar(self.ResponseText)
        self.ScrollBar.pack(side=tk.RIGHT, fill=tk.Y)
        self.ScrollBar.config(command=self.ResponseText.yview)
        # self.ResponseText.config(yscrollcommand=self.ScrollBar.set)

        self.test_profile_change("")
        self.test_file_generate()

    def antenna_change(self, event):
        if self.AntennaNum.get() == "2":
            self.SpatialStream.config(state=tk.NORMAL)
            self.SpatialStream.delete(0, tk.END)
            self.SpatialStream.insert(tk.END, "1,2")
            if self.TestProfile.get() == "Full":
                self.SpatialStream.config(state=tk.DISABLED)
        else:
            self.SpatialStream.config(state=tk.NORMAL)
            self.SpatialStream.delete(0, tk.END)
            self.SpatialStream.insert(tk.END, "1,2,3,4")
            if self.TestProfile.get() == "Full":
                self.SpatialStream.config(state=tk.DISABLED)

    def test_profile_change(self, event):
        if self.TestProfile.get() == "Full":
            self.Direction.config(state=tk.DISABLED)
            self.AC_ChannelList.config(state=tk.DISABLED)
            self.N_ChannelList.config(state=tk.DISABLED)
            self.ClientNum.config(state=tk.DISABLED)
            self.TrafficType.config(state=tk.DISABLED)
            self.SpatialStream.config(state=tk.DISABLED)
            self.AC_MCS.config(state=tk.DISABLED)
            self.N_MCS.config(state=tk.DISABLED)
            self.N_Bandwidth.config(state=tk.DISABLED)
            self.FrameSize.config(state=tk.DISABLED)
        else:
            self.ClientNum.config(state=tk.NORMAL)
            self.Direction.config(state=tk.NORMAL)
            self.TrafficType.config(state=tk.NORMAL)
            self.FrameSize.config(state=tk.NORMAL)

            if self.TestType.get()=="802.11ac" or self.TestType.get()=="802.11ax_5G":
                self.AC_ChannelList.config(state=tk.NORMAL)
                self.AC_MCS.config(state=tk.NORMAL)
                self.SpatialStream.config(state=tk.NORMAL)
                self.N_MCS.config(state=tk.DISABLED)
                self.N_Bandwidth.config(state=tk.DISABLED)
                self.N_ChannelList.config(state=tk.DISABLED)
            else:
                self.AC_ChannelList.config(state=tk.DISABLED)
                self.AC_MCS.config(state=tk.DISABLED)
                self.SpatialStream.config(state=tk.DISABLED)
                self.N_MCS.config(state=tk.NORMAL)
                self.N_Bandwidth.config(state=tk.NORMAL)
                self.N_ChannelList.config(state=tk.NORMAL)

    def start_test(self):
        global vChassis_ip_addr
        global vLicense_ip_addr
        global vPort_AC
        global vPort_N
        global vPort_ETH
        global vLog_Dir
        global vModel
        global vVersion
        global vTest_Profile
        global vFrame
        global vProject_Type

        global report_generate
        logging.info("Start Veriwave Automation Testing")
        report_generate = True

        vChassis_ip_addr = self.ChassisIP.get()
        vLicense_ip_addr = self.LicenseIP.get()
        vPort_AC = self.AC_Port.get()
        vPort_N = self.N_Port.get()
        vPort_ETH = self.EthPort.get()
        vLog_Dir = self.LogDir.get()
        vModel = self.Model.get()
        vVersion = self.Version.get()

        vTest_Profile = self.TestProfile.get()
        vTest_Type = self.TestType.get()
        vProject_Type = self.ProjectType.get()

        vClient_Num = self.ClientNum.get()
        vChannel_AC = self.AC_ChannelList.get()
        vChannel_N = self.N_ChannelList.get()
        vMcs_ac = self.AC_MCS.get()
        vMcs_n = self.N_MCS.get()
        vBandwidth_n = self.N_Bandwidth.get()
        vStream = self.SpatialStream.get()
        vDirection = self.Direction.get()
        vTraffic_Type = self.TrafficType.get()
        vFrame = self.FrameSize.get()

        if self.ProjectType.get()=="HGU":
            vCPE_Console = CPE_CONSOLE
            vQTN_Console = QTN_CONSOLE

            self.set_thread = set_cpe2(vQTN_Console, b"\r\niwpriv wifi0_0 set_vopt 0\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\ncall_qcsapi set_optim_stats wifi0_0 1\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\niwpriv wifi0_0 set_vopt 1\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\ncall_qcsapi set_optim_stats wifi0_0 0\r\n")
            self.after(1000, "")
            self.set_thread.join()

        else:
            vCPE_Console = CPE_CONSOLE
            vQTN_Console = CPE_CONSOLE
            self.set_thread = set_cpe2(vQTN_Console, b"\r\nsh\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\niwpriv wifi0_0 set_vopt 0\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\ncall_qcsapi set_optim_stats wifi0_0 1\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\niwpriv wifi0_0 set_vopt 1\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\ncall_qcsapi set_optim_stats wifi0_0 0\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(vQTN_Console, b"\r\nexit\r\n")
            self.after(1000, "")
            self.set_thread.join()

        logging.info("Configure CPE/QTN board for Veriwave Testing")

        #generate fw_name
        today = date.today()
        report_name = 'veriwave_report' + "_" + vModel + "_" + vVersion + "_" + today.strftime("%Y%m%d") + ".xlsx"
        #report_name = 'veriwave_report' + "_" + model + "_" + version + ".xlsx"

        for iClient in vClient_Num.split(","):
            if vTest_Type == "802.11ac" or vTest_Type=="802.11ax_5G":
                for iChannel in vChannel_AC.split(","):
                    for iMcs in vMcs_ac.split(","):
                        for iStream in vStream.split(","):
                            for iDirection in vDirection.split(","):
                                if iDirection == "Unidirectional":
                                    for iTraffic_Type in vTraffic_Type.split(","):
                                        for iFlow in FLOW_DIR:
                                            if iFlow == "TX":
                                                vTraffic_Src = "eth_group"
                                                vTraffic_Dst = "ac_group"
                                            elif iFlow == "RX":
                                                vTraffic_Src = "ac_group"
                                                vTraffic_Dst = "eth_group"
                                            self.run_testcase(vTest_Type, iClient, iChannel, iMcs, "80", iStream, iDirection, iTraffic_Type, iFlow, vTraffic_Src, vTraffic_Dst, vProject_Type, report_name)

                                elif iDirection == "Bidirectional":
                                    for iTraffic_Type in vTraffic_Type.split(","):
                                        vTraffic_Src = "eth_group"
                                        vTraffic_Dst = "ac_group"
                                        iFlow = "BI"
                                        if iFlow =="BI" and iTraffic_Type =="Tcp":
                                            logging.info("Ignore TCP Bi-directional testing")
                                        else:
                                            self.run_testcase(vTest_Type, iClient, iChannel, iMcs, "80", iStream, iDirection, iTraffic_Type, iFlow, vTraffic_Src, vTraffic_Dst, vProject_Type, report_name)

            elif vTest_Type == "802.11n" or vTest_Type=="802.11ax_2.4G":
                for iChannel in vChannel_N.split(","):
                    for iBandwidth in vBandwidth_n.split(","):
                        for iMcs in vMcs_n.split(","):
                            #for iStream in vStream.split(","):
                            if iMcs == "7":
                                iStream = "1"
                            elif iMcs == "15":
                                iStream = "2"
                            elif iMcs == "23":
                                iStream = "3"
                            elif iMcs == "31":
                                iStream = "4"

                            for iDirection in vDirection.split(","):
                                if iDirection == "Unidirectional":
                                    for iTraffic_Type in vTraffic_Type.split(","):
                                        for iFlow in FLOW_DIR:
                                            if iFlow == "TX":
                                                vTraffic_Src = "eth_group"
                                                vTraffic_Dst = "n_group"
                                            elif iFlow == "RX":
                                                vTraffic_Src = "n_group"
                                                vTraffic_Dst = "eth_group"
                                            self.run_testcase(vTest_Type, iClient, iChannel, iMcs, iBandwidth, iStream, iDirection, iTraffic_Type, iFlow, vTraffic_Src, vTraffic_Dst, vProject_Type, report_name)

                                elif iDirection == "Bidirectional":
                                    for iTraffic_Type in vTraffic_Type.split(","):
                                        vTraffic_Src = "eth_group"
                                        vTraffic_Dst = "n_group"
                                        iFlow = "BI"
                                        if iFlow == "BI" and iTraffic_Type == "Tcp":
                                            logging.info("Ignore TCP Bi-directional testing")
                                        else:
                                            self.run_testcase(vTest_Type, iClient, iChannel, iMcs, iBandwidth, iStream, iDirection, iTraffic_Type, iFlow, vTraffic_Src, vTraffic_Dst, vProject_Type, report_name)
        tk.messagebox.showinfo("INFO", "Test Finished")

    def run_testcase(self, test_type, client, channel, mcs, bandwidth, stream, direction, traffic_type, flow, src, dst, project, report_name):
        global vChassis_ip_addr
        global vLicense_ip_addr
        global vPort_AC
        global vPort_N
        global vPort_ETH
        global vLog_Dir
        global vModel
        global vVersion
        global vTest_Profile
        global vFrame
        global vQTN_Console
        global vCPE_Console

        script_file = "PQA_" + test_type + "_" + traffic_type.upper() + ".tcl"
        test_program = "tclsh vw_auto.tcl -f " + script_file + " "
        if test_type =="802.11ac" or test_type=="802.11ax_5G":
            band = "5G"
        elif test_type =="802.11n" or test_type=="802.11ax_2.4G":
            band = "2.4G"

        if project == "HGU":
            prefix = "tefapp "
        else:
            prefix = ""

        if flow != "BI":
            test_folder = vLog_Dir + "\\\\auto_" + band + "_" + "C" + client + "_CH" + channel + "_MCS" + mcs + "_BW" + bandwidth + "_S" + stream + "_" + direction \
                      + "_" + traffic_type.upper() + "_" + flow + "\\\\"
        else:
            test_folder = vLog_Dir + "\\\\auto_" + band + "_" + "C" + client + "_CH" + channel + "_MCS" + mcs + "_BW" + bandwidth + "_S" + stream + "_" + direction \
                      + "_" + traffic_type.upper() + "_BI\\\\"

        if test_type=="802.11ac" or test_type=="802.11ax_5G":
            # --addLicServer " + vLicense_ip_addr
            if mcs=="0":
                vw_mcs = "9"
            else:
                vw_mcs = mcs
            test_case = "--noconfig --debug 3 --var ip_addr " + vChassis_ip_addr + " " + \
                    "--var ac_port " + vPort_AC + " --var n_port " + vPort_N + " --var eth_port " + vPort_ETH + " --var logdir " + test_folder + " " + \
                    "--var direction " + direction + " --var channel_ac " + channel + " --var channel_n 1" + " --var channel " + channel + " " + \
                    "--var client_num " + client + " --var src " + src + " --var dst " + dst + " " + \
                    "--var traffic " + traffic_type + " --var stream " + stream + " --var mcs_ac " + vw_mcs + " --var mcs_n 31 " + "--var bandwidth_n 40 " + \
                    "--var frame \"" + vFrame.replace(","," ") + "\" --var version " + vVersion + " --var model " + vModel + " --var flow " + traffic_type.upper()

        elif test_type=="802.11n" or test_type=="802.11ax_2.4G":
            test_case = "--noconfig --debug 3 --var ip_addr " + vChassis_ip_addr + " " + \
                        "--var ac_port " + vPort_AC + " --var n_port " + vPort_N + " --var eth_port " + vPort_ETH + " --var logdir " + test_folder + " " + \
                        "--var direction " + direction + " --var channel_ac 36" + " --var channel_n " + channel + " --var channel " + channel + " " + \
                        "--var client_num " + client + " --var src " + src + " --var dst " + dst + " " + \
                        "--var traffic " + traffic_type + " --var stream " + stream + " --var mcs_ac 9" + " --var mcs_n " + mcs + " " + \
                        "--var bandwidth_n " + bandwidth + " --var frame \"" + vFrame.replace(","," ") + "\" --var version " + vVersion + " --var model " + vModel + \
                        " --var flow " + traffic_type.upper()

        self.ResponseText.insert(tk.END, "Test Folder: " + test_folder + "\r\n")
        self.ResponseText.insert(tk.END, "Test Parameter: " + test_program + test_case + "\r\n")
        self.ResponseText.insert(tk.END, "\r\n")
        self.ResponseText.insert(tk.END, "Start Configure CPE..................\r\n")

        logging.info("Configure CPE Channel Info")
        if test_type == "802.11ac" or test_type=="802.11ax_5G":
            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi_plus channel " + channel.encode('utf_8') + b"\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi_plus mcs " + mcs.encode('utf_8') + b"\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi_plus nss " + stream.encode('utf_8') + b"\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi_plus save\r\n")
            self.after(1000, "")
            self.set_thread.join()

        elif test_type == "802.11n" or test_type=="802.11ax_2.4G":
            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi channel " + channel.encode('utf_8') + b"\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi mcs " + mcs.encode('utf_8') + b"\r\n")
            self.after(1000, "")
            self.set_thread.join()

            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi nss " + stream.encode('utf_8') + b"\r\n")
            self.after(1000, "")
            self.set_thread.join()

            if bandwidth == "20":
                self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi bandwidth 20\r\n")
                self.after(1000, "")
                self.set_thread.join()
            else:
                self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi bandwidth auto\r\n")
                self.after(1000, "")
                self.set_thread.join()

            self.set_thread = set_cpe2(CPE_CONSOLE, b"\r\n" + prefix.encode() + b"set wifi save\r\n")
            self.after(1000, "")
            self.set_thread.join()

        self.ResponseText.focus_set()
        self.ResponseText.focus_lastfor()
        self.ResponseText.see(tk.END)
        self.update()
        self.update_idletasks()
        app.update()

        self.ResponseText.insert(tk.END, "Waiting for CPE's WiFi for 10s.......\r\n")
        logging.info("Waiting CPE for 10s to bring up WiFi interface")
        for count in range(0, 10):
            self.after(1000, "")
            app.update()
            app.update_idletasks()

        self.ResponseText.insert(tk.END, "Running Veriwave with test case.......\r\n")

        logging.info("Running vw_auto.tcl")
        #For testing comand
        #self.thread = exec_command("start /wait ping 8.8.8.8 -n 60")
        self.thread = exec_command("start /wait " + test_program + test_case)
        while self.thread.is_alive():
            self.after(100,"")
            self.update()
            self.update_idletasks()
            app.update()

        #self.poll_thread(self.thread)
        #self.thread.join()
        test_folder = test_folder.replace("\\\\", "\\")

        #if self.check_file(test_folder)==True:
        logging.info("Parser JSON file and write data to EXCEL")
        logging.info("Data Folder: " + test_folder)

        if mcs=="9":
            self.parser_json_result(vModel, vVersion, test_type, test_folder, 0, False, report_name)
        elif mcs=="0":
            self.parser_json_result(vModel, vVersion, test_type, test_folder, 0, True, report_name)
        else:
            self.parser_json_result(vModel, vVersion, test_type, test_folder, 0, False, report_name)

        #else:
        #    print("Parser JSON fail, Please check test script\r\n")

        return

    def poll_thread(self, work_thread):
        logging.debug('[DEBUG] thread checking: %s', work_thread)
        #state = self.StartButton['state']
        #if self.proc_state == False and state != tk.NORMAL:
        #if self.proc_state == False:
        #    self.proc_state = True
        #    self.thread = exec_command(msg)

        if work_thread.is_alive():
            self.update()
            self.update_idletasks()
            app.update()
            self.alive_id = self.after(1000, lambda: self.poll_thread(work_thread))
        else:
            self.update()
            self.update_idletasks()
            app.update()
            #self.proc_state = False
            #self.stop_id = self.after(1000, lambda: self.poll_thread(work_thread))

    def run_cmd(self, cmd, callback=None, watch=False):
        """Runs the given command and gathers the output.
        If a callback is provided, then the output is sent to it, otherwise it
        is just returned.
        Optionally, the output of the command can be "watched" and whenever new
        output is detected, it will be sent to the given `callback`.
        Args:
            cmd (str): The command to run.
        Kwargs:
            callback (func):  The callback to send the output to.
            watch (bool):     Whether to watch the output of the command.  If True,
                              then `callback` is required.
        Returns:
            A string containing the output of the command, or None if a `callback`
            was given.
        Raises:
            RuntimeError: When `watch` is True, but no callback is given.
        """
        if watch and not callback:
            raise RuntimeError('You must provide a callback when watching a process.')

        output = None
        try:
            proc = subprocess.Popen(shlex.split(cmd), stdout=subprocess.PIPE)

            if watch:
                while proc.poll() is None:
                    line = proc.stdout.readline()
                    if line != "":
                        callback(line)

                # Sometimes the process exits before we have all of the output, so
                # we need to gather the remainder of the output.
                remainder = proc.communicate()[0]
                if remainder:
                    callback(remainder)
            else:
                output = proc.communicate()[0]
        except:
            err = str(sys.exc_info()[1]) + "\n"
            output = err

        if callback and output is not None:
            callback(output)
            return None

        return output

    def flush(self, message):
        sys.stdout.write(message)
        print(message)
        sys.stdout.flush()

    def generate(self):
        global report_generate
        vLog_Dir = self.LogDir.get()
        vTest_Type = self.TestType.get()
        vModel = self.Model.get()
        vVersion = self.Version.get()
        today = date.today()

        report_generate=True
        report_name = 'veriwave_report' + "_" + vModel + "_" + vVersion + "_" + today.strftime("%Y%m%d") + ".xlsx"
        self.parser_json_result(vModel, vVersion, vTest_Type,vLog_Dir, 0, True, report_name)
        # self.parser_json_result("802.11n","veriwave_report.xlsx",vLo, 0, True)
        tk.messagebox.showinfo("INFO","Generate Finished")

        # report = self.TemplateText.get()
        # html = self.HTMLText.get()
        # raw_data = self.RawText.get()
        # tester = self.tester_radio.get()
        # mcs9_auto = self.channel_check.get()
        #
        # if report != '':
        #     if html != '' or raw_data != '':
        #         if html != '':
        #             self.parser_html_result(report, html)
        #             msgbox.showinfo("Information", "Convert HTML finished!")
        #         if raw_data != '':
        #             self.parser_json_result(report, raw_data, tester, mcs9_auto)
        #             msgbox.showinfo("Information", "Convert JSON finished!")
        #     else:
        #         msgbox.showerror("Error", "Please input either HTML file or Raw data folder")
        # else:
        #     msgbox.showerror("Error", "Template file is required!!")
        return

    def check_file(self,report_folder):
        for file in os.listdir(report_folder):
            if file.endswith(".json"):
                logging.info(os.path.join(report_folder, file))
                return True
        return False

    def test_file_generate(self):

        test_type_value = {"802.11ax_5G_default": "keylset ac_group EnableAMSDUtxaggregation True",
            "802.11ax_2.4G_default": "keylset n_group EnableAMSDUtxaggregation False",
            "802.11ac_default": "keylset ac_group EnableAMSDUtxaggregation True",
            "802.11n_default": "keylset n_group EnableAMSDUtxaggregation False",
            "802.11ax_5G": "keylset ac_group EnableAMSDUtxaggregation True",
            "802.11ax_2.4G": "keylset n_group EnableAMSDUtxaggregation True",
            "802.11ac": "keylset ac_group EnableAMSDUtxaggregation False",
            "802.11n": "keylset n_group EnableAMSDUtxaggregation False"}

        # default:
        traffic_type_value = {"default": ["keylset global_config TrialDuration 10","keylset global_config TestDurationSec 10"],
          "Udp": ["keylset global_config TrialDuration 10", "keylset global_config TestDurationSec 10"],
          "Tcp": ["keylset global_config TrialDuration 30","keylset global_config TestDurationSec 30"]}

        for test_type in TEST_TYPE:
            for traffic_type in TRAFFIC_TYPE:
                # Read in the file
                with open("PQA.tcl", 'r') as file:
                    filedata = file.read()
                # Replace the target string
                filedata = filedata.replace(test_type_value[test_type + "_default"], test_type_value[test_type])
                count = 0
                for keylet in traffic_type_value["default"]:
                    filedata = filedata.replace(keylet, traffic_type_value[traffic_type][count])
                    count += 1
                # Write the file out again
                with open("PQA_" + test_type + "_" + traffic_type.upper() + ".tcl", 'w') as file:
                    file.write(filedata)

    def parser_json_result(self, model, version, band, rawdata_folder, tester_id, mcs9_to_auto, fw_name):
        global report_generate
        cell_col = 0
        cell_row = 0
        row_count = 1
        col_count = 1

        title = ['Folder', 'Model', 'FW version', 'Channel', 'Bandwidth', 'MCS', 'Antenna', 'Pattern_Flow Direction',
                 '128', '256', '512', '1024', '1280', '1518']

        if report_generate == True:
            copyfile("veriwave_throughput_template.xlsx", fw_name)

        try:
            workbook = openpyxl.load_workbook(fw_name)
        except Exception as e:
            logging.info("Open Excel error!")
            logging.info(repr(e))

        if band == "802.11ac" or band=="802.11ax_5G":
            worksheet_report = workbook[band]
            if report_generate == True:
                worksheet_raw = workbook.create_sheet('5G_raw_data' + "_" + band)
                self.raw_row_count = row_count
            else:
                worksheet_raw = workbook['5G_raw_data' + "_" + band]
        elif band == "802.11n" or band=="802.11ax_2.4G":
            worksheet_report = workbook[band]
            if report_generate == True:
                worksheet_raw = workbook.create_sheet('2.4G_raw_data')
                self.raw_row_count = row_count
            else:
                worksheet_raw = workbook['2.4G_raw_data']

        if report_generate == True:
            for name in title:
                worksheet_raw.cell(self.raw_row_count, col_count, name)
                col_count = col_count + 1
            self.raw_row_count += 1

        report_generate = False

        col_count = 1
        row_count = row_count + 1

        for dirPath, dirNames, fileNames in os.walk(rawdata_folder):
            # for file in fileNames:
            if os.path.exists(dirPath + '\\' + 'Results_unicast_throughput.json') == True:
                logging.info(dirPath)
                #worksheet_raw.cell(row_count, 1, dirPath)
                table = json.load(open(dirPath + '\\' + 'Results_unicast_throughput.json'))

                model = table['Config']['DutInfo']['AP Model']
                sw_name = table['Config']['DutInfo']['AP SW Version']

                for port in table['Config']['Ports']:
                    for id_name in table['Config']['Ports'][port]:
                        if id_name == 'Channel':
                            channel = 'CANAL' + str(table['Config']['Ports'][port][id_name])


                for group in table['Config']['ClientGroups']:
                    for setting in table['Config']['ClientGroups'][group]:
                        if setting == 'acPhySettings':
                            MCS_rate = 'MCS_' + str(table['Config']['ClientGroups'][group][setting]['VhtDataMcs'])
                            ant_num = str(table['Config']['ClientGroups'][group][setting]['NumSpatialStreams'])
                            bandwidth = "80"
                        elif setting == "nPhySettings":
                            MCS_rate = 'MCS_' + str(table['Config']['ClientGroups'][group][setting]['DataMcsIndex'])
                            bandwidth = 'BW' + str(table['Config']['ClientGroups'][group][setting]['ChannelBandwidth'])
                            #ant_num = str(table['Config']['ClientGroups'][group][setting]['NumSpatialStreams'])
                            ant_num = ""
                            if MCS_rate == "MCS_7":
                                ant_num = "1"
                            elif MCS_rate == "MCS_15":
                                ant_num = "2"
                            elif MCS_rate == "MCS_23":
                                ant_num = "3"
                            elif MCS_rate == "MCS_31":
                                ant_num = "4"

                if mcs9_to_auto == True and MCS_rate == 'MCS_9':
                    MCS_rate = 'MCS_0'

                if "_MCS9_" in dirPath:
                    MCS_rate = 'MCS_9'
                elif "_MCS0_" in dirPath:
                    MCS_rate = 'MCS_0'

                ant_num = ant_num + 'x' + ant_num
                pat_type = str(table['Config']['Traffics']['TrafficType']).upper()

                dir_type = str(table['Config']['Mapping']['Map'])
                if dir_type == 'Ethernet to Wireless':
                    dir_type = 'DOWN'
                else:
                    dir_type = 'UP'
                uni_type = str(table['Config']['Mapping']['Direction'])
                if uni_type != 'Unidirectional':
                    dir_type = 'BI'

                pat_type = pat_type + '_' + dir_type
                logging.info(model + "," + version  + "," + str(channel)  + "," + str(bandwidth)  + "," + str(MCS_rate)  + "," + ant_num  + "," +  pat_type  + "," +  dir_type  + "," + uni_type)

                worksheet_raw.cell(self.raw_row_count, 1, dirPath)
                worksheet_raw.cell(self.raw_row_count, 2, model)
                worksheet_raw.cell(self.raw_row_count, 3, version)
                worksheet_raw.cell(self.raw_row_count, 4, channel)
                worksheet_raw.cell(self.raw_row_count, 5, bandwidth)
                worksheet_raw.cell(self.raw_row_count, 6, MCS_rate)
                worksheet_raw.cell(self.raw_row_count, 7, ant_num)
                worksheet_raw.cell(self.raw_row_count, 8, pat_type)
                col_count = 9
                error_ignore = False
                try:
                    if band == "802.11ac" or band=="802.11ax_5G":
                        cell_row = ac_data_loc[channel] + ac_data_loc[pat_type]
                    elif band == "802.11n" or band=="802.11ax_2.4G":
                        cell_row = n_data_loc[channel] + n_data_loc[pat_type]
                    if tester_id == '1':
                        if band == "802.11ac" or band=="802.11ax_5G":
                            cell_col = ac_data_loc[MCS_rate] + ac_data_loc[ant_num] + AC_TEF_LOC_COL
                    else:
                        if band == "802.11ac" or band=="802.11ax_5G":
                            cell_col = ac_data_loc[MCS_rate] + ac_data_loc[ant_num]
                        elif band == "802.11n" or band=="802.11ax_2.4G":
                            cell_col = n_data_loc[bandwidth] + n_data_loc[ant_num]
                    error_ignore = False
                except Exception as e:
                    logging.info('Raw Data Parser Exception: ' + repr(e))
                    error_ignore = True

                try:
                    shift_row = 0
                    for tput_data in table['Results']['Summary Results']:
                        frame_size = str(tput_data['FrameSize'])
                        pkt_data = round(tput_data['Throughputbps'] / 1000 / 1000, 2)
                        if band == "802.11ac" or band=="802.11ax_5G":
                            shift_row = cell_row + ac_data_loc[frame_size]
                        elif band == "802.11n" or band=="802.11ax_2.4G":
                            shift_row = cell_row + n_data_loc[frame_size]
                        if error_ignore == False:
                            worksheet_report.cell(shift_row, cell_col, pkt_data)
                        worksheet_raw.cell(self.raw_row_count, col_count, pkt_data)
                        col_count = col_count + 1
                except Exception as e:
                    logging.info('Raw Data Parser Exception: ' + repr(e))

                row_count = row_count + 1
                self.raw_row_count += 1
            elif os.path.exists(dirPath + '\\' + 'Results_unicast_throughput.csv') == True:
                # <.*>(.*)<\/.*>
                # <Channel>36</Channel>
                # <VhtDataMcs>9</VhtDataMcs>
                # <NumSpatialStreams>'4'</NumSpatialStreams>
                # <Type>UDP</Type>
                # <MappingOptions>0</MappingOptions>
                # <FlowDirection>Unidirectional</FlowDirection>
                data_csv = self.csv_read(dirPath + '\\' + 'Results_unicast_throughput.csv')
                data_wml = self.csv_read(dirPath + '\\' + 'test_unicast_unidirectional_throughput.wml')

                #if '115034' in dirPath:
                #    print('pause')

                result = {}
                data_start = bool
                for line in data_csv:
                    if 'AP Model' in line:
                        result['model'] = line.split(',')[1]
                    elif 'AP SW Version' in line:
                        result['sw_name'] = line.split(',')[1]
                    elif 'Frame Size' in line:
                        data_start = True
                    elif data_start == True:
                        size = line.split(',')[0]
                        throughput = float(line.split(',')[7]) / 1000 / 1000
                        result[size] = throughput
                for line in data_wml:
                    content = ''
                    regex = re.compile("<.*>(.*)<\/.*>")
                    match = regex.search(line)
                    if bool(match) == True:
                        content = match.group(1)
                    if '<Channel>' in line and content != 'NA':
                        result['channel'] = 'CANAL' + content
                    elif '<NumSpatialStreams>' in line and content != 'NA':
                        content = content.replace('\'', '')
                        result['Antenna'] = content + 'x' + content
                    elif '<VhtDataMcs>' in line and content != 'NA':
                        result['MCS'] = 'MCS_' + content
                    elif '<MappingOptions>' in line and content != 'NA':
                        if content == '0':
                            result['direction'] = 'DOWN'
                        elif content == '1':
                            result['direction'] = 'UP'
                    elif '<FlowDirection>' in line and content != 'NA':
                        result['uni-bi'] = content
                    elif '<Type>' in line and content != 'NA':
                        result['Type'] = content
                        if result['uni-bi'] == 'Unidirectional':
                            result['Type'] = result['Type'] + '_' + result['direction']
                        else:
                            result['Type'] = Result['Type'] + '_BI'

                if mcs9_to_auto == True and result['MCS'] == 'MCS_9':
                    result['MCS'] = 'MCS_0'

                worksheet_raw.cell(row_count, 1, dirPath)
                worksheet_raw.cell(row_count, 2, result['model'])
                worksheet_raw.cell(row_count, 3, result['sw_name'])
                worksheet_raw.cell(row_count, 4, result['channel'])
                worksheet_raw.cell(row_count, 5, result['MCS'])
                worksheet_raw.cell(row_count, 6, result['Antenna'])
                worksheet_raw.cell(row_count, 7, result['Type'])
                col_count = 8
                try:
                    cell_row = ac_data_loc[result['channel']] + ac_data_loc[result['Type']]
                    if tester_id == '1':
                        cell_col = ac_data_loc[result['MCS']] + ac_data_loc[result['Antenna']] + AC_TEF_LOC_COL
                    else:
                        cell_col = ac_data_loc[result['MCS']] + ac_data_loc[result['Antenna']]
                    error_ignore = False
                except Exception as e:
                    logging.info('Exception:' + repr(e))
                    error_ignore = True

                try:
                    shift_row = 0
                    for frame_size in result:
                        if frame_size == '128' or frame_size == '256' or frame_size == '512' \
                                or frame_size == '1024' or frame_size == '1280' or frame_size == '1518':
                            pkt_data = round(result[frame_size], 2)
                            shift_row = cell_row + ac_data_loc[frame_size]
                            if error_ignore == False:
                                worksheet_report.cell(shift_row, cell_col, pkt_data)
                            worksheet_raw.cell(row_count, col_count, pkt_data)
                            col_count = col_count + 1
                except Exception as e:
                    logging.info('Exception: ' + repr(e))
                col_count = 1
                row_count = row_count + 1
                #print(result)

        workbook.save(fw_name)
        workbook.close()
        return 0

    def set_cpe(self, title, command):
        try:
            hwndMain = win32gui.FindWindow("VTWin32", title)
            # hwndMain = win32gui.FindWindow("ConsoleWindowClass",None)
            #print(hwndMain)
            win32gui.ShowWindow(hwndMain, win32con.SW_MAXIMIZE)
            win32gui.SetForegroundWindow(hwndMain)
            win32gui.SetActiveWindow(hwndMain)


            # win32gui.SetFocus(hwndMain)
            time.sleep(2)
            hwndChild = win32gui.GetWindow(hwndMain, win32con.GW_CHILD)
            #send_enter()
            #typer(command)
            #send_enter()

            hwndMain = win32gui.FindWindow("TkTopLevel", "Veriwave Automation")
            # hwndMain = win32gui.FindWindow("ConsoleWindowClass",None)
            #print(hwndMain)
            win32gui.SetForegroundWindow(hwndMain)
            win32gui.SetActiveWindow(hwndMain)

        except Exception as e:
            logging.info(repr(e))

    def csv_read(self, filename):
        table = []
        logging.info('open filename = ' + filename)
        file = open(filename, mode="r", errors='ignore')
        table = file.readlines()
        return table

    def start_logging(self):
        # Enable the logging to console and file
        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        logging.basicConfig(level=output_logging,
                            format='%(asctime)s: [%(levelname)s] %(message)s',
                            datefmt='%a, %d %b %Y %H:%M:%S',
                            filename='veriwave_test.log',
                            filemode='w')

        console = logging.StreamHandler()
        console.setLevel(output_logging)
        formatter = logging.Formatter('%(levelname)-4s %(message)s')
        console.setFormatter(formatter)
        logging.getLogger('').addHandler(console)

class COPYDATASTRUCT(ctypes.Structure):
    _fields_ = [
        ('dwData', ctypes.wintypes.LPARAM),
        ('cbData', ctypes.wintypes.DWORD),
        ('lpData', ctypes.c_char_p)
        #formally lpData is c_void_p, but we do it this way for convenience
    ]

class exec_command(Thread):
    def __init__(self, cmd):
        super().__init__()
        self.task = 0
        self.cmd = cmd
        self.daemon = True
        self.start()

    def run(self):
        logging.info("Execute WaveAutomation command: " + self.cmd)
        devnull = open(os.devnull, 'wb')
        args = shlex.split(self.cmd)

        os.system(self.cmd)
        #si = subprocess.STARTUPINFO()
        #si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        #self.task = subprocess.Popen(args, shell=True, startupinfo=si, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        #                             stdin=subprocess.PIPE)
        #self.task.wait()

        #for line in io.TextIOWrapper(self.task.stdout, encoding="utf-8"):
        #    logging.info(line)
        #self.task.terminate()

class set_cpe2(Thread):
    def __init__(self, title, msg):
        super().__init__()
        self.task = 0
        self.title = title
        self.msg = msg
        self.daemon = True
        self.start()

    def run(self):
        # prepare data and send it
        FindWindow = ctypes.windll.user32.FindWindowW
        SendMessage = ctypes.windll.user32.SendMessageW
        try:
            hwnd = FindWindow("VTWin32", self.title)
            cds = COPYDATASTRUCT()
            cds.dwData = 1
            cds.cbData = ctypes.sizeof(ctypes.create_string_buffer(self.msg))
            cds.lpData = ctypes.c_char_p(self.msg)
            SendMessage(hwnd, win32con.WM_COPYDATA, 0, ctypes.byref(cds))
        except Exception as e:
            logging.info("Write command to console failed, " + repr(e))

if __name__ == '__main__':
    app = Auto_Veriwave()
    app.mainloop()
    sys.exit(0)
